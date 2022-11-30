from platform import python_version
import os
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side
from datetime import datetime
from openpyxl.utils import get_column_letter as gs
import streamlit as st
start_time = datetime.now()

# Help





def octant_analysis(mod=5000):
    pass

# Read all the excel files in a batch format from the input\ folder. Only xlsx to be allowed
# Save all the excel files in a the output\ folder. Only xlsx to be allowed
# output filename = input_filename[_octant_analysis_mod_5000].xlsx , ie, append _octant_analysis_mod_5000 to the original filename.


# Code
ver = python_version()

if ver == "3.8.10":
    print("Correct Version Installed")
else:
    print("Please install 3.8.10. Instruction are present in the GitHub Repo\Webmail. Url: https:\pastebin.com\nvibxmjw")

def rowtraversal(rowi, sheet):
    mx = -1e9
    val = -1
    for i in range(8):
        valt = sheet.cell(row=rowi, column=36+i).value
        if(valt>mx):
            mx = valt
            val = 36+i
            
    sheet.cell(row=rowi, column = val).fill = PatternFill("solid", start_color="FFFF00")    

thin_border = Border(left = Side(style='thin'),right = Side(style='thin'),top= Side(style='thin'),bottom = Side(style='thin'))

def decide_octant(u, v, w):
    # Function to decide the octant on the basis of given values of u, v, and w
    if (u > 0 and v > 0 and w > 0):
        return "1"
    elif (u > 0 and v > 0 and w < 0):
        return "-1"
    elif (u < 0 and v < 0 and w > 0):
        return "3"
    elif (u < 0 and v < 0 and w < 0):
        return "-3"
    elif (u < 0 and v > 0 and w > 0):
        return "2"
    elif (u < 0 and v > 0 and w < 0):
        return "-2"
    elif (u > 0 and v < 0 and w > 0):
        return "4"
    elif (u > 0 and v < 0 and w < 0):
        return "-4"


PathA = os.getcwd()

# os.makedirs("output", exist_ok=True)

mod = 5000
MOD = mod

octants = ["1", "-1", "2", "-2", "3", "-3", "4", "-4"]

dict = {"1": 0, "-1": 0, "2": 0, "-2": 0, "3": 0, "-3": 0, "4": 0, "-4": 0}
octantnames = ["Internal outward interaction",
               "External outward interaction",
               "External Ejection",
               "Internal Ejection",
               "External inward interaction",
               "Internal inward interaction",
               "Internal sweep",
               "External sweep"
               ]
# octants and their names
dict4octants = {}
for i in octants:
    dict4octants[i] = 0


def outpathfun(xlfile, MOD):
    fname = xlfile[0:-5]
    fname += "octant_analysis_mod"+(str)(MOD)+".xlsx"
    # print(fname)
    # print(fname)
    return fname
# returning the name of output file

dict4first = {}


def FunToSort(rowi, sheet, list1, octants, keynames, dict4first):
    # function to print the sorted index on the sheet
    list = []
    for i in range(8):
        list.append([sheet.cell(row=rowi, column=15+i).value, i+1])
    list.sort(reverse=True)
    # appended the values on a list and then reverse sorted it
    for i in range(8):
        # print(list[i][1])
        list1.append(list[i][1])
    for i in range(8):
        sheet.cell(row=rowi, column=22+list1[i]).value = i+1
        sheet.cell(row=rowi, column=22+list1[i]).border = thin_border
        if (i+1 == 1):
            sheet.cell(row=rowi, column=22 +
                       list1[i]).fill = PatternFill("solid", start_color="FFFF00")
    sheet.cell(row=rowi, column=31).value = octants[list1[0]-1]
    sheet.cell(row=rowi, column=31).border = thin_border
    dict4first[octants[list1[0]-1]] += 1
    sheet.cell(row=rowi, column=32).value = keynames[list1[0]-1]
    sheet.cell(row=rowi, column=32).border = thin_border
    # finally printed the values on the sheet


def mainfun(file):
    # main function which will be executed for each of the file in input
    # wb = load_workbook(f"{PathA}\input\{file}")
    wb = file
    print(type(file))
    # print(f"{PathA}\input\{file}")
    sheet = wb
    sheet.insert_rows(1)
    sheet.cell(row=1, column=14).value = "Overall Octant Count"
    sheet.cell(row=1, column=35).value = "Overall Transition Count"
    sheet.cell(row=1, column=45).value = "Longest Subsequence Length"
    sheet.cell(row=1, column=49).value = "Longest Subsequence Length with Range"
    headings = ["T", "U", "V", "W", "U Avg", "V Avg", "W Avg",
                "U'=U-U avg", "V'=V-V avg", "W'=W-W avg", "Octant"]

    for i in range(len(headings)):
        sheet.cell(row=2, column=1+i).value = headings[i]

    totalrow = sheet.max_row
    totalcols = sheet.max_column

    list = [0, 0, 0]
    for i in range(3):
        for j in range(2, totalrow):
            list[i] += float(sheet.cell(row=j+1, column=i+2).value)

    list2 = []
    for i in list:
        y = i/(totalrow-1)
        yt = y
        x = round(y, 3)
        temp = (str)(x)
        if temp[0] == '-' and temp[1] == '0':
            yt = 0

        list2.append(format(yt, ".3f"))

    # print(list2)
    for i in range(len(list2)):
        sheet.cell(row=3, column=5+i).value = list2[i]
        # sheet.cell(row=3, column=5+i).border = thin_border

    for i in range(2, totalrow):
        a = sheet.cell(row=i+1, column=2).value
        b = sheet.cell(row=i+1, column=3).value
        c = sheet.cell(row=i+1, column=4).value
        sheet.cell(row=i+1, column=8).value = format(float(a)-float(list2[0]),".3f")
        sheet.cell(row=i+1, column=9).value = format(float(b)-float(list2[1]),".3f")
        sheet.cell(row=i+1, column=10).value = format(float(c)-float(list2[2]),".3f")
        d = float(a)-float(list2[0])
        e = float(b)-float(list2[1])
        f = float(c)-float(list2[2])
        sheet.cell(row=i+1, column=11).value = decide_octant(d, e, f)

        # Completed till octant determination

        dict4octants[decide_octant(d, e, f)] += 1

    sheet.cell(row=3, column=14).value = "Octant ID"
    sheet.cell(row=3, column=14).border = thin_border
    sheet.cell(row=4, column=13).value = "Mod " + (str)(MOD)
    for i in range(len(octants)):
        sheet.cell(row=3, column=15+i).value = octants[i]
        sheet.cell(row=3, column=15+i).border = thin_border
        strtemp = "Rank of " + str(octants[i])
        sheet.cell(row=3, column=23+i).value = strtemp
        sheet.cell(row=3, column=23+i).border = thin_border
        
        
    sheet.cell(row=3, column=31).value = "Rank 1 Octant ID"
    sheet.cell(row=3, column=31).border = thin_border
    sheet.cell(row=3, column=32).value = "Rank 1 Octant Name"
    sheet.cell(row=3, column=32).border = thin_border
    sheet.cell(row=4, column=14).value = "Overall Count"
    sheet.cell(row=4, column=14).border = thin_border

    sheet.cell(row=12, column=27).value = "Octant ID"
    sheet.cell(row=12, column=27).border = thin_border
    sheet.cell(row=12, column=28).value = "Octant Name"
    sheet.cell(row=12, column=28).border = thin_border
    sheet.cell(row=12, column=29).value = "Count of Rank 1 Mod Values"
    sheet.cell(row=12, column=29).border = thin_border
    
    for i in range(8):
        sheet.cell(row=13+i, column=27).value = octants[i]
        sheet.cell(row=13+i, column=27).border = thin_border
        sheet.cell(row=13+i, column=28).value = octantnames[i]
        sheet.cell(row=13+i, column=28).border = thin_border
    
    listtemp = []
    dicttemp = {}
    for i in octants:
        dicttemp[i] = 0

    for i in octants:
        dict4first[i] = 0

    for i in range(len(octants)):
        sheet.cell(row=4, column=15+i).value = dict4octants[octants[i]]
        sheet.cell(row=4, column=15+i).border = thin_border
    
    FunToSort(4, sheet, listtemp, octants, octantnames, dicttemp)

    workable = totalrow - 1
    divs = workable//MOD
    divs += 1
    if (workable % MOD == 0):
        divs -= 1

    res = 0
    # listoflastvals = [2]
    for i in range(divs):
        rlv = res*MOD
        rhv = ((res+1)*MOD)-1
        rhv = min(rhv, totalrow-2)
        # listoflastvals.append(rhv)
        sheet.cell(row=5+i, column=14).value = str(rlv) + "-" + str(rhv)
        sheet.cell(row=5+i, column=14).border = thin_border
        res += 1

    dict4ranges = {}
    for i in octants:
        dict4ranges[i] = 0

    rown = 1
    valn = 0
    listtemp2 = []

    for i in range(2, totalrow):
        if (rown % MOD == 0):
            listtemp2.clear()
            ocn = 0
            for oc in dict4ranges:
                sheet.cell(row=5+valn, column=15+ocn).value = dict4ranges[octants[ocn]]
                sheet.cell(row=5+valn, column=15+ocn).border = thin_border
                ocn += 1
            FunToSort(5+valn, sheet, listtemp2,
                      octants, octantnames, dict4ranges)
            # dict4ranges = {}
            # print(dict4ranges)
            dict4ranges.clear()
            for j in octants:
                dict4ranges[j] = 0
            valn += 1
        # print(type(i))
        dict4ranges[sheet.cell(row=1+i, column=11).value] += 1
        rown += 1

    FunToSort(5+valn, sheet, listtemp2, octants, octantnames, dict4ranges)
    valn += 1
    listtemp2.clear()

    # dict4ranges = {}
    ocn = 0
    for oc in dict4ranges:
        sheet.cell(row=4+valn, column=15+ocn).value = dict4ranges[octants[ocn]]
        sheet.cell(row=4+valn, column=15+ocn).border = thin_border
        ocn += 1
    # print(dict4ranges)
    dict4ranges.clear()
    for j in octants:
        dict4ranges[j] = 0

    for i in range(divs):
        dict4first[sheet.cell(row=4+i, column=31).value] += 1

    for i in range(8):
        sheet.cell(row=13+i, column=29).value = dict4first[octants[i]]
        sheet.cell(row=13+i, column=29).border = thin_border

    listtemp2 = ["Octant ###","Longest Subsequence Length","Count"]
    for i in range(3):
        sheet.cell(row=3, column=49+i).value = listtemp2[i]
        sheet.cell(row=3, column=49+i).border = thin_border
        
    listtemp3 = ["Octant ##", "Longest Subsequence Length", "Count"]
    for i in range(len(listtemp3)):
        sheet.cell(row=3, column=45+i).value = listtemp3[i]
        sheet.cell(row=3, column=45+i).border = thin_border

    # sheet.insert_cols(50)
    prev = str("1")
    prcnt = 0
    for i in range(3, totalrow+1):
        k = str(sheet.cell(row=i, column=11).value)
        # print(k,prev)
        if k is not prev:
            # print(k)
            temp = max(dict[prev], prcnt)
            dict[prev] = temp
            prcnt = 1
            prev = k
        else:
            prcnt += 1
    dict[prev] = max(dict[prev], prcnt)

    frq_of_maxm_subs = {"1": 0, "-1": 0, "2": 0,
                        "-2": 0, "3": 0, "-3": 0, "4": 0, "-4": 0}
    # dictionary to store the frequency of maximum length of subsequence

    starting_points = {"1": [], "-1": [], "2": [],
                       "-2": [], "3": [], "-3": [], "4": [], "-4": []}
    # dictionary to store the starting point times of maximum length subsequence
    ending_points = {"1": [], "-1": [], "2": [],
                     "-2": [], "3": [], "-3": [], "4": [], "-4": []}
    prev = "1"
    prcnt = 0
    stp = sheet.cell(row=2, column=1).value
    etp = sheet.cell(row=2, column=1).value
    for i in range(3, totalrow+1):
        k = str(sheet.cell(row=i, column=11).value)
        # print(k,prev)
        if k is not prev:
            # print(k)
            if prcnt == dict[prev]:
                frq_of_maxm_subs[prev] += 1
                # increasing the count of maximum length subsequence
                starting_points[prev].append(stp)
                ending_points[prev].append(etp)
                # storing the starting and ending point of the subsequences
            prcnt = 1
            prev = k
            stp = sheet.cell(row=i, column=1).value
        else:
            prcnt += 1
        etp = sheet.cell(row=i, column=1).value
    # print(frq_of_maxm_subs)

    # print(starting_points)
    # print(ending_points)

    # printing the subsequence time range values.
    k = 4

    for i in range(8):
        sheet.cell(row=k, column=50-1).border = thin_border
        sheet.cell(row=k, column=50-1).value = octants[i]
        sheet.cell(row=k, column=51-1).border = thin_border
        sheet.cell(row=k, column=51-1).value = dict[octants[i]]
        sheet.cell(row=k, column=52-1).border = thin_border
        sheet.cell(row=k, column=52-1).value = frq_of_maxm_subs[octants[i]]
        k += 1
        sheet.cell(row=k, column=50-1).border = thin_border
        sheet.cell(row=k, column=50-1).value = "Time"
        sheet.cell(row=k, column=51-1).border = thin_border
        sheet.cell(row=k, column=51-1).value = "From"
        sheet.cell(row=k, column=52-1).border = thin_border
        sheet.cell(row=k, column=52-1).value = "To"
        k += 1
        for j in range(frq_of_maxm_subs[octants[i]]):
            sheet.cell(row=k, column=51-1).border = thin_border
            sheet.cell(row=k, column=51-1).value = starting_points[octants[i]][j]
            sheet.cell(row=k, column=52-3).border = thin_border
            sheet.cell(row=k, column=52-1).border = thin_border
            sheet.cell(row=k, column=52-1).value = ending_points[octants[i]][j]
            k += 1

    for i in range(8):
        sheet.cell(row=4+i, column=47).border = thin_border
        sheet.cell(row=4+i, column=47).value = frq_of_maxm_subs[octants[i]]
        sheet.cell(row=4+i, column=45).border = thin_border
        sheet.cell(row=4+i, column=45).value = octants[i]
        sheet.cell(row=4+i, column=46).border = thin_border
        sheet.cell(row=4+i, column=46).value = dict[octants[i]]

    dict2 = {}
    tx = 0
    for i in octants:
        dict2[i] = tx
        tx+=1

    lastval = 1
    sheet.cell(row=lastval+1, column=36).value = "To"
    sheet.cell(row=lastval+3, column=34).value = "From"
    sheet.cell(row=lastval+2, column=35).value = "Octant #"
    sheet.cell(row=lastval+2, column=35).border = thin_border
    for i in range(len(octants)):
        sheet.cell(row=lastval+2, column=36+i).border = thin_border
        sheet.cell(row=lastval+2, column=36+i).value = octants[i]
        sheet.cell(row=lastval+3+i, column=35).border = thin_border
        sheet.cell(row=lastval+3+i, column=35).value = octants[i]

    modcnt = totalrow//MOD
    modcnt += 1
    if totalrow % MOD == 0:
        modcnt -= 1

    lv = 0
    hv = MOD-1

    constr = lastval+3

    strtvals = []

    for i in range(modcnt):
        lastval += 14
        sheet.cell(row=lastval+1, column=35).value = str(lv)+" - "+str(hv)
        sheet.cell(row=lastval, column=35).value = "Mod Transition Count"
        sheet.cell(row=lastval+1, column=36).value = "To"
        sheet.cell(row=lastval+2, column=35).value = "Count"
        sheet.cell(row=lastval+2, column=35).border = thin_border
        sheet.cell(row=lastval+3, column=34).value = "From"
        strtvals.append(lastval+3)
        lv = hv+1
        hv += MOD
        hv = min(hv, totalrow)

        for i in range(8):
            sheet.cell(row=lastval+2, column=36+i).border = thin_border
            sheet.cell(row=lastval+2, column=36+i).value = octants[i]
        for i in range(8):
            sheet.cell(row=lastval+3+i, column=35).border = thin_border
            sheet.cell(row=lastval+3+i, column=35).value = octants[i]

    k = 0
    cnt = 0
    prev = sheet.cell(row=3, column=11).value
    # print(strtvals)
    # print(prev)
    for i in range(4, totalrow):
        curr = sheet.cell(row=i, column=11).value
        tempprev = dict2[prev]
        tempcurr = dict2[curr]
        temp1 = sheet.cell(row=constr+tempprev, column=36+tempcurr).value
        if (temp1 != None):
            sheet.cell(row=constr+tempprev, column=36 +tempcurr).border = thin_border
            sheet.cell(row=constr+tempprev, column=36 +tempcurr).value = int(temp1)+1
        else:
            sheet.cell(row=constr+tempprev, column=36+tempcurr).border = thin_border
            sheet.cell(row=constr+tempprev, column=36+tempcurr).value = 1

        temp = sheet.cell(
            row=strtvals[k]+tempprev, column=36+tempcurr).value
        if (temp != None):
            sheet.cell(row=strtvals[k]+tempprev,column=36+tempcurr).border = thin_border
            sheet.cell(row=strtvals[k]+tempprev,column=36+tempcurr).value = int(temp)+1
        else:
            sheet.cell(row=strtvals[k]+tempprev,column=36+tempcurr).border = thin_border
            sheet.cell(row=strtvals[k]+tempprev,column=36+tempcurr).value = 1

        cnt += 1
        prev = str(curr)
        if cnt % MOD == 0:
            k += 1
# finally filled the octant tables for different mod ranges
    strtvals.insert(0, 4)
    # print(strtvals)
    for val in strtvals:
        for i in range(8):
            for j in range(8):
                if (sheet.cell(row=val+i, column=36+j).value == None):
                    sheet.cell(row=val+i, column=36+j).value = 0
                        
    for i in strtvals:
        for j in range(8):
            rowtraversal(i+j, sheet)
            

    opname = outpathfun(file, MOD)
    for i in range(60):
        sheet.column_dimensions[gs(i+1)].width = 25
    wb.save(f"{PathA}\output\{opname}")
    print(f"{opname}, file processed.")
    # print(f"{PathA}\output\{opname}")
    


# octant_analysis(mod)
# dir_list = os.listdir(f"{PathA}\input")
# # print(dir_list)

os.chdir(PathA)
os.makedirs("output", exist_ok=True)

dir_list = st.file_uploader("Choose a xlsx file",accept_multiple_files=True)

for i in dir_list:
    mainfun(i)
# This shall be the last lines of the code.
end_time = datetime.now()
print('Duration of Program Execution: {}'.format(end_time - start_time))